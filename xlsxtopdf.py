"""
Script untuk convert file Excel (banyak sheet) menjadi PDF terpisah per sheet.

CARA KERJA:
1. Baca file Excel asli.
2. Untuk tiap sheet, bikin file Excel sementara yang isinya cuma 1 sheet itu.
3. Convert file sementara itu ke PDF pakai LibreOffice (command line, headless).
4. Hasil PDF disimpan di folder output, dengan nama sesuai nama sheet.

REQUIREMENT:
- Python library: openpyxl  ->  pip install openpyxl
- LibreOffice harus terinstall di komputer (soffice ada di PATH).
  - Windows: install LibreOffice biasa, lalu cari lokasi soffice.exe
    (biasanya: C:\\Program Files\\LibreOffice\\program\\soffice.exe)
  - Linux/Mac: sudo apt install libreoffice  (atau brew install libreoffice)

CARA PAKAI:
    python excel_to_pdf_per_sheet.py "nama_file.xlsx" "folder_output"

Kalau folder_output tidak diisi, default-nya folder "output_pdf" di lokasi script ini.
"""

import sys
import os
import shutil
import subprocess
import tempfile
import re
import openpyxl
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter


def setup_page_fit(ws):
    """
    Atur page setup sheet supaya:
    - Orientasi landscape (lebih lebar, muat lebih banyak kolom)
    - Fit-to-width 1 halaman (kolom tidak kepotong ke halaman berikutnya)
    - Fit-to-height dibiarkan mengalir ke bawah (baris boleh lanjut ke halaman baru)
    - Margin dikecilin biar lebih banyak ruang
    - Print area di-set ke seluruh area yang ada isinya
    """
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 0 = tidak dibatasi tinggi, tinggal lanjut ke halaman baru
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins = PageMargins(
        left=0.3, right=0.3, top=0.4, bottom=0.4, header=0.2, footer=0.2
    )

    if ws.max_row and ws.max_column:
        last_col_letter = get_column_letter(ws.max_column)
        ws.print_area = f"A1:{last_col_letter}{ws.max_row}"


def sanitize_filename(name: str) -> str:
    """Bersihin nama sheet dari karakter yang gak boleh dipakai di nama file."""
    name = re.sub(r'[\\/*?:"<>|]', "_", name)
    return name.strip() or "sheet"


def find_soffice():
    """Cari lokasi executable LibreOffice (soffice)."""
    candidates = [
        "soffice",
        "soffice.exe",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        "/usr/bin/soffice",
        "/opt/libreoffice/program/soffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]
    for c in candidates:
        if shutil.which(c):
            return shutil.which(c)
        if os.path.isfile(c):
            return c
    return None


def convert_sheet_to_pdf(sheet_xlsx_path: str, output_dir: str, soffice_path: str):
    """Convert 1 file xlsx (isi 1 sheet) jadi PDF pakai LibreOffice headless."""
    cmd = [
        soffice_path,
        "--headless",
        "--convert-to", "pdf",
        "--outdir", output_dir,
        sheet_xlsx_path,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"  [GAGAL] {os.path.basename(sheet_xlsx_path)} -> {result.stderr.strip()}")
        return False
    return True


def main():
    if len(sys.argv) < 2:
        print("Cara pakai: python excel_to_pdf_per_sheet.py <file_excel.xlsx> [folder_output]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.getcwd(), "output_pdf")

    if not os.path.isfile(input_path):
        print(f"File tidak ditemukan: {input_path}")
        sys.exit(1)

    soffice_path = find_soffice()
    if not soffice_path:
        print("LibreOffice (soffice) tidak ditemukan di komputer ini.")
        print("Install dulu LibreOffice, atau edit variabel 'candidates' di fungsi find_soffice().")
        sys.exit(1)

    os.makedirs(output_dir, exist_ok=True)

    print(f"Membuka file: {input_path}")
    wb = openpyxl.load_workbook(input_path, data_only=True)

    sheet_names = wb.sheetnames
    print(f"Ditemukan {len(sheet_names)} sheet: {sheet_names}")

    with tempfile.TemporaryDirectory() as tmpdir:
        for sheet_name in sheet_names:
            print(f"\nMemproses sheet: {sheet_name}")

            # Bikin workbook baru yang isinya cuma 1 sheet ini
            single_wb = openpyxl.load_workbook(input_path, data_only=True)
            for name in list(single_wb.sheetnames):
                if name != sheet_name:
                    single_wb.remove(single_wb[name])

            setup_page_fit(single_wb[sheet_name])

            safe_name = sanitize_filename(sheet_name)
            temp_xlsx_path = os.path.join(tmpdir, f"{safe_name}.xlsx")
            single_wb.save(temp_xlsx_path)

            ok = convert_sheet_to_pdf(temp_xlsx_path, tmpdir, soffice_path)
            if not ok:
                continue

            # File PDF hasil convert akan bernama sama dengan xlsx-nya
            generated_pdf = os.path.join(tmpdir, f"{safe_name}.pdf")
            final_pdf = os.path.join(output_dir, f"{safe_name}.pdf")

            if os.path.isfile(generated_pdf):
                shutil.move(generated_pdf, final_pdf)
                print(f"  -> Berhasil: {final_pdf}")
            else:
                print(f"  [GAGAL] PDF tidak ditemukan untuk sheet '{sheet_name}'")

    print(f"\nSelesai! Cek hasil PDF di folder: {output_dir}")


if __name__ == "__main__":
    main()